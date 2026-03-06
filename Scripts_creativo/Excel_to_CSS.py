from openpyxl import Workbook,load_workbook
from pathlib import Path

ARQUIVO_A_CONVERTER = Path(__file__).parent / 'planilha_maior.xlsx'
SAIDA= Path(__file__).parent/'planilha_convertida.html'

wb = load_workbook(ARQUIVO_A_CONVERTER)
ws = wb.active

i=0
linhas = ws.max_row
colunas = ws.max_column

print('debugando excel_to_CSS.py')


head= '''
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CSS GRID</title>
</head>
'''

Has_header = 0
while True:
    header_input= input('A tabela possui cabeçalho? (Y/n)')
    if header_input.upper() == 'Y':
        Has_header = True 
        break
    elif header_input.lower() == 'n':
        Has_header = False
        break
    else:
        print(print('resposta inválida, digite apenas (Y) ou (n)'))


def create_style(ws):
    linhas = ws.max_row
    colunas = ws.max_column
    style = '<style>\n'
    style += '''.grid {display: grid;
grid-template-columns:'''
    if Has_header:
        i=0
        for row in ws.iter_rows():
            for cell in row:
                if i==0:
                    print(cell)
                    value = cell.value.replace(' ','_')
                    style+= f'[{value}] 100px '
            i+=1
    else:
        for i in range(colunas):
            style+='100px '
    style+=';\n'
    style+='}'

    style+='\n .grid div {border: 0.5px solid rgb(199, 199, 199)}'

    style+='\n </style>'
    print(style)
    return style
style = create_style(ws)



def create_body(ws):
    body = '''
    <div class=grid> 
    '''
    for row in ws.iter_rows():
        for sheet in row:
            body+= f'<div id={sheet.coordinate}>{sheet.value}</div>\n'
    body+='</div>'
    return body 
body = create_body(ws)
    



print("Arquivo será salvo em:", SAIDA.resolve())


with open(str(SAIDA), 'w') as file:
    file.write(head)
    file.write(style)
    file.write(body)
print('linha final executada')

print(ws.cell(row=1,column=1).value)