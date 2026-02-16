from openpyxl import Workbook
import random
from datetime import date, timedelta
import os
caminho = os.path.join(os.path.expanduser("~"), "Desktop", "Prova_Excel.xlsx")


wb = Workbook()

# ======================
# ABA 1 — ALUNOS
# ======================
ws_alunos = wb.active
ws_alunos.title = "Alunos"

ws_alunos.append(["Aluno", "Nota 1", "Nota 2", "Nota 3", "Média", "Situação"])

nomes_alunos = [
    "Ana", "Bruno", "Carla", "Daniel", "Eduarda", "Felipe", "Gabriela",
    "Henrique", "Isabela", "João", "Karina", "Lucas", "Mariana",
    "Nicolas", "Otávio", "Paula", "Rafael", "Sofia", "Thiago", "Vitória"
]

for i in range(1, 101):  # 100 alunos
    nome = random.choice(nomes_alunos) + f" {i}"
    notas = [round(random.uniform(3, 10), 1) for _ in range(3)]
    ws_alunos.append([nome, *notas, "", ""])

# ======================
# ABA 2 — PRODUTOS
# ======================
ws_prod = wb.create_sheet("Produtos")
ws_prod.append(["Código", "Produto", "Preço"])

produtos = [
    (101, "Mouse", 50),
    (102, "Teclado", 120),
    (103, "Monitor", 800),
    (104, "Headset", 200),
    (105, "Notebook", 3500),
    (106, "Webcam", 300),
]

for p in produtos:
    ws_prod.append(p)

# ======================
# ABA 3 — VENDAS
# ======================
ws_vendas = wb.create_sheet("Vendas")
ws_vendas.append([
    "Data", "Vendedor", "Código Produto", "Produto",
    "Quantidade", "Preço Unitário", "Valor Total"
])

vendedores = ["João", "Maria", "Pedro", "Lucas", "Ana"]
codigos_produtos = [p[0] for p in produtos]

data_inicial = date(2026, 1, 1)

for _ in range(200):  # 200 vendas
    data = data_inicial + timedelta(days=random.randint(0, 30))
    vendedor = random.choice(vendedores)
    codigo = random.choice(codigos_produtos)

    preco = next(p[2] for p in produtos if p[0] == codigo)
    quantidade = random.randint(1, 10)

    ws_vendas.append([
        data.strftime("%d/%m/%Y"),
        vendedor,
        codigo,
        "",
        quantidade,
        preco,
        ""
    ])

# ======================
# ABA 4 — FUNCIONÁRIOS
# ======================
ws_func = wb.create_sheet("Funcionários")
ws_func.append([
    "Nome", "Sobrenome", "Setor", "Salário", "Nome Completo"
])

nomes = ["Lucas", "Mariana", "Rafael", "Paula", "Carlos", "Fernanda", "Igor"]
sobrenomes = ["Silva", "Souza", "Oliveira", "Lima", "Pereira", "Costa"]
setores = ["Marketing", "Financeiro", "TI", "RH", "Vendas"]

for _ in range(50):  # 50 funcionários
    nome = random.choice(nomes)
    sobrenome = random.choice(sobrenomes)
    setor = random.choice(setores)
    salario = random.randint(1500, 5000)

    ws_func.append([nome, sobrenome, setor, salario, ""])

# ======================
# SALVAR
# ======================
wb.save(r"C:\Users\sarge\OneDrive\Documentos\ProvaExcelGrande")
print("Planilha criada com sucesso!")

print("Arquivo salvo em:", os.getcwd())
input("Pressione ENTER para fechar...")